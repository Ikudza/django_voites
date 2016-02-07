# -*- coding: utf-8 -*-
"""
Initialization celery as applications and initialization methods to tasks
"""
from __future__ import unicode_literals
from clients.models import Client
import os
from celery import Celery
from django.conf import settings

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "voite_project.settings")
celery_app = Celery("voite_project")
celery_app.config_from_object("django.conf:settings")
celery_app.autodiscover_tasks(lambda: settings.INSTALLED_APPS)


@celery_app.task
def create_list_clients_xlsx(response):
    """
        Create report list clients to xlsx
    """
    from openpyxl import Workbook

    fields = ['name', 'surname', 'age', 'birthday', 'voite']
    wb = Workbook()
    ws = wb.active
    ws.title = u"Clients"
    clients = Client.objects.all()
    for i in xrange(len(clients)):
        for j in xrange(1, len(fields)+1):
            if i == 0:
                ws.cell(row=i+1, column=j).value = fields[j-1]
            ws.cell(row=i+2, column=j).value = clients[i][fields[j-1]]
    wb.save(response)
    return response
